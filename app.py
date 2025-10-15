# app.py — SDoH Bilingual Survey (Streamlit)
# - Bilingual (English/Español)
# - 100+ questions with branching (adaptive follow-ups)
# - All visible questions shown in order, numbered
# - Radios start with NO default selection (index=None; Streamlit >= 1.25)
# - Exports: per-submission JSON, cumulative CSV/XLSX (no pandas)

import streamlit as st
import json, csv, glob, datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
import pandas as pd

st.sidebar.title("Navigation")
page = st.sidebar.radio("Go to", ["Survey", "Download Responses"])

# Password-protected Download Page
if page == "Download Responses":
    st.title("📥 Secure Download — Survey Responses")

    # ---- Step 1: Set your password ----
    # Option A: Hardcode (simple)
    PASSWORD = "Health2025"

    # Option B (recommended): Load from environment variable
    # import os
    # PASSWORD = os.getenv("SDOH_DOWNLOAD_PASS", "Health2025")

    # ---- Step 2: Ask for password ----
    st.markdown("Please enter the access password to view and download survey data:")
    pw_input = st.text_input("Password", type="password")

    if pw_input == "":
        st.info("🔒 Enter password to proceed.")
        st.stop()

    elif pw_input != PASSWORD:
        st.error("❌ Incorrect password.")
        st.stop()

    else:
        st.success("✅ Access granted.")
        st.write("You can now download the collected responses below:")

        csv_path = "sdh_responses.csv"
        xlsx_path = "sdh_responses.xlsx"

        # ---- Step 3: Offer secure downloads ----
        from pathlib import Path
        if Path(csv_path).exists():
            with open(csv_path, "rb") as f:
                st.download_button("⬇️ Download CSV", f, file_name="sdh_responses.csv", mime="text/csv")
        else:
            st.warning("No CSV file found yet.")

        if Path(xlsx_path).exists():
            with open(xlsx_path, "rb") as f:
                st.download_button("⬇️ Download Excel (XLSX)", f, file_name="sdh_responses.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            st.warning("No Excel file found yet.")

    st.stop()


if page == "Download Responses":
    st.title("📥 Download Collected Responses")

    csv_path = "sdh_responses.csv"
    xlsx_path = "sdh_responses.xlsx"

    # Ensure file exists
    if Path(csv_path).exists():
        with open(csv_path, "rb") as f:
            st.download_button("⬇️ Download CSV", f, file_name="sdh_responses.csv", mime="text/csv")
    else:
        st.warning("No CSV responses found yet.")

    if Path(xlsx_path).exists():
        with open(xlsx_path, "rb") as f:
            st.download_button("⬇️ Download Excel (XLSX)", f, file_name="sdh_responses.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        st.warning("No Excel responses found yet.")

    st.stop()


# =========================
# Helpers & File Writers
# =========================
def now_iso():
    return datetime.datetime.now().isoformat(timespec="seconds")

def flatten_for_table(record, prefer_labels=True):
    flat = {}
    def _f(prefix, obj):
        if isinstance(obj, dict):
            if prefer_labels and set(obj.keys()) == {"code", "label"}:
                flat[prefix[:-1]] = obj.get("label")
                return
            for k, v in obj.items():
                _f(f"{prefix}{k}.", v)
        elif isinstance(obj, list):
            for i, v in enumerate(obj):
                _f(f"{prefix}{i}.", v)
        else:
            flat[prefix[:-1]] = obj
    _f("", record)
    return flat

def write_csv(rows, cols, csv_path):
    Path(csv_path).parent.mkdir(parents=True, exist_ok=True)
    with open(csv_path, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for r in rows:
            out = {k: r.get(k, "") for k in cols}
            w.writerow(out)

def write_xlsx(rows, cols, xlsx_path, sheet_name="responses"):
    Path(xlsx_path).parent.mkdir(parents=True, exist_ok=True)
    if Path(xlsx_path).exists():
        wb = load_workbook(xlsx_path)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
    ws.append(cols)
    for r in rows:
        ws.append([r.get(c, "") for c in cols])
    # autosize-ish
    for col_idx, col_name in enumerate(cols, start=1):
        width = max(10, min(60, int(max([len(str(col_name))] + [len(str(r.get(col_name,""))) for r in rows]) * 1.05)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
    wb.save(xlsx_path)

def save_all_outputs(record, csv_path="sdh_responses.csv", xlsx_path="sdh_responses.xlsx"):
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    json_path = f"sdh_response_{ts}.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(record, f, indent=2, ensure_ascii=False)

    # rebuild CSV/XLSX from all JSONs
    json_files = sorted(glob.glob("sdh_response_*.json"))
    rows, all_keys = [], set(["timestamp"])
    for jf in json_files:
        with open(jf, "r", encoding="utf-8") as f:
            rec = json.load(f)
        flat = flatten_for_table(rec, prefer_labels=True)
        row = {"timestamp": rec.get("meta", {}).get("completed_at")}
        row.update(flat)
        rows.append(row)
        all_keys.update(row.keys())
    cols = ["timestamp"] + sorted(k for k in all_keys if k != "timestamp")
    write_csv(rows, cols, csv_path)
    write_xlsx(rows, cols, xlsx_path, sheet_name="responses")
    return json_path, csv_path, xlsx_path

# =========================
# Question Bank (bilingual + branching)
# =========================
def opt(en, es, code=None): 
    return {"code": code, "en": en, "es": es}

# Common option sets
YN = [opt("Yes", "Sí", 1), opt("No", "No", 2)]
YN_DK = [opt("Yes","Sí",1), opt("No","No",2), opt("Don't know","No sabe",9), opt("Prefer not to answer","Prefiere no responder",7)]
FREQ5 = [opt("Always","Siempre",1), opt("Often","A menudo",2), opt("Sometimes","A veces",3), opt("Rarely","Rara vez",4), opt("Never","Nunca",5)]
HL_CONF = [opt("Extremely","Extremadamente",1), opt("Quite a bit","Bastante",2), opt("Somewhat","Algo",3), opt("A little bit","Un poco",4), opt("Not at all","Nada",5)]
GEN_HEALTH = [opt("Excellent","Excelente",1), opt("Very good","Muy buena",2), opt("Good","Buena",3), opt("Fair","Regular",4), opt("Poor","Mala",5)]
ENG = [opt("Very well","Muy bien",1), opt("Well","Bien",2), opt("Not well","No muy bien",3), opt("Not at all","Nada",4)]
INCOME = [
    opt("< $10,000","< $10,000",1), opt("$10,000–$24,999","$10,000–$24,999",2), opt("$25,000–$34,999","$25,000–$34,999",3),
    opt("$35,000–$49,999","$35,000–$49,999",4), opt("$50,000–$74,999","$50,000–$74,999",5),
    opt("$75,000–$99,999","$75,000–$99,999",6), opt("$100,000–$149,999","$100,000–$149,999",7),
    opt("$150,000–$199,999","$150,000–$199,999",8), opt("$200,000 or more","$200,000 o más",9),
    opt("Prefer not to answer","Prefiere no responder",10)
]
EDU = [
    opt("Less than high school","Menos que secundaria",1), opt("High school or GED","Secundaria o GED",2),
    opt("Some college, no degree","Algo de universidad, sin título",3), opt("Associate degree","Técnico/Asociado",4),
    opt("Bachelor’s degree","Licenciatura",5), opt("Master’s degree","Maestría",6), opt("Professional/Doctoral degree","Profesional/Doctorado",7)
]
EMP = [
    opt("Working now","Trabajando actualmente",1), opt("Temporary leave","Permiso temporal",2),
    opt("Looking for work / Unemployed","Buscando trabajo / Desempleado(a)",3), opt("Retired","Jubilado(a)",4),
    opt("Disabled","Con discapacidad",5), opt("Keeping house","Labores del hogar",6), opt("Student","Estudiante",7), opt("Other","Otro",8)
]
MAR = [
    opt("Married","Casado(a)",1), opt("Divorced","Divorciado(a)",2), opt("Widowed","Viudo(a)",3),
    opt("Separated","Separado(a)",4), opt("Never married","Nunca casado(a)",5), opt("Member of an unmarried couple","Pareja no casada",6)
]
SO1 = [
    opt("Gay","Gay",1), opt("Lesbian","Lesbiana",2), opt("Straight (not gay/lesbian)","Heterosexual (no gay/lesbiana)",3),
    opt("Bisexual","Bisexual",4), opt("None of these — show more options","Ninguna de estas — ver más opciones",5)
]
SO2 = [
    opt("Queer","Queer",1), opt("Poly/omni/sapio/pansexual","Poli/omni/sapio/pansexual",2),
    opt("Asexual","Asexual",3), opt("Two-spirit","Dos espíritus",4), opt("Figuring it out","En proceso de definir",5),
    opt("Mostly straight","Mayormente heterosexual",6), opt("No sexuality","No se considera con sexualidad",7),
    opt("No labels","No usa etiquetas",8), opt("Don't know","No sabe",9),
    opt("Something else (specify)","Otra cosa (especifique)",10), opt("Prefer not to answer","Prefiere no responder",11)
]
FS_OSN = [opt("Often true","A menudo cierto",1), opt("Sometimes true","A veces cierto",2), opt("Never true","Nunca cierto",3)]
FS_YN = [opt("Yes","Sí",1), opt("No","No",2)]
FS_3A = [opt("Almost every month","Casi todos los meses",1), opt("Some months but not every month","Algunos meses pero no todos",2), opt("Only 1 or 2 months","Solo 1 o 2 meses",3)]
AGREE5 = [opt("Strongly agree","Totalmente de acuerdo",1), opt("Agree","De acuerdo",2), opt("Neutral","Neutral",3), opt("Disagree","En desacuerdo",4), opt("Strongly disagree","Totalmente en desacuerdo",5)]

QUESTIONS = []
def add_q(section, _id, en, es, qtype, options=None, branch=None):
    QUESTIONS.append({
        "id": _id, "section": section,
        "text": {"en": en, "es": es},
        "type": qtype,
        "options": options or [],
        "branch": branch
    })

# ---------- Access to Health Services (~10) ----------
add_q("Access to Health Services","q1_last_visit_any",
      "How long has it been since you last saw a doctor or other health professional about your health?",
      "¿Cuánto tiempo ha pasado desde la última vez que vio a un médico u otro profesional de la salud por su salud?",
      "radio",
      options=[
          opt("Within the past year (<12 months)","Dentro del último año (<12 meses)",1),
          opt("Within the last 2 years","En los últimos 2 años",2),
          opt("Within the last 3 years","En los últimos 3 años",3),
          opt("Within the last 5 years","En los últimos 5 años",4),
          opt("Within the last 10 years","En los últimos 10 años",5),
          opt("10 years ago or more","Hace 10 años o más",6),
          opt("Never","Nunca",0),
          opt("Prefer not to answer","Prefiere no responder",7),
          opt("Don't know","No sabe",9),
      ])
add_q("Access to Health Services","q2_last_visit_wellness",
      "Was that visit a wellness/physical/general check-up?",
      "¿Esa visita fue un examen de bienestar/físico/revisión general?",
      "radio", options=YN_DK,
      branch=lambda a: a.get("q1_last_visit_any",{}).get("code") in (1,2,3,4,5,6))
add_q("Access to Health Services","q3_last_wellness",
      "How long has it been since your last wellness/physical/general check-up?",
      "¿Cuánto tiempo ha pasado desde su último examen de bienestar/físico/revisión general?",
      "radio",
      options=[
          opt("Within the past year","Dentro del último año",1),
          opt("Within the last 2 years","En los últimos 2 años",2),
          opt("Within the last 3 years","En los últimos 3 años",3),
          opt("Within the last 5 years","En los últimos 5 años",4),
          opt("More than 5 years","Más de 5 años",5),
          opt("Prefer not to answer","Prefiere no responder",7),
          opt("Don't know","No sabe",9),
      ],
      branch=lambda a: a.get("q1_last_visit_any",{}).get("code") in (1,2,3,4,5,6) and a.get("q2_last_visit_wellness",{}).get("code") not in (None,1))
add_q("Access to Health Services","q4_usual_source",
      "Is there a place you USUALLY go if you're sick and need care?",
      "¿Tiene un lugar al que USUALMENTE va cuando está enfermo y necesita atención?",
      "radio",
      options=[
          opt("Yes","Sí",1), opt("There is NO place","NO hay lugar",2),
          opt("There is MORE THAN ONE place","Hay MÁS DE UN lugar",3),
          opt("Prefer not to answer","Prefiere no responder",7),
          opt("Don't know","No sabe",9),
      ])
add_q("Access to Health Services","q5_usual_place_type",
      "What kind of place is it/do you go to most often?",
      "¿Qué tipo de lugar es o al que va con más frecuencia?",
      "radio",
      options=[
          opt("Doctor's office / health center","Consultorio / centro de salud",1),
          opt("Walk-in / urgent care / retail clinic","Clínica sin cita / urgencias / minorista",2),
          opt("Emergency room","Sala de emergencias",3),
          opt("VA Medical Center / VA clinic","Centro médico/Clínica de VA",4),
          opt("Some other place","Otro lugar",5),
          opt("Does not go to one place most often","No va a un solo lugar con más frecuencia",6),
          opt("Prefer not to answer","Prefiere no responder",7),
          opt("Don't know","No sabe",9),
      ],
      branch=lambda a: a.get("q4_usual_source",{}).get("code") in (1,3))
add_q("Access to Health Services","q6_urgent_care_visits",
      "In the past 12 months, how many times did you go to urgent care or a retail clinic? (0–96; 96 = 96+)",
      "En los últimos 12 meses, ¿cuántas veces fue a urgencias o a una clínica minorista? (0–96; 96 = 96+)",
      "int")
add_q("Access to Health Services","q7_er_visits",
      "In the past 12 months, how many times did you go to a hospital emergency room? (0–96; 96 = 96+)",
      "En los últimos 12 meses, ¿cuántas veces fue a la sala de emergencias de un hospital? (0–96; 96 = 96+)",
      "int")
add_q("Access to Health Services","q8_hospitalized_overnight",
      "In the past 12 months, were you hospitalized overnight?",
      "En los últimos 12 meses, ¿estuvo hospitalizado durante la noche?",
      "radio", options=YN_DK)
add_q("Access to Health Services","q9_delayed_care_cost",
      "In the past 12 months, did you DELAY medical care because of cost?",
      "En los últimos 12 meses, ¿RETRASÓ la atención médica por el costo?",
      "radio", options=YN_DK)
add_q("Access to Health Services","q10_unmet_need_cost",
      "In the past 12 months, was there a time you NEEDED care but DID NOT GET IT because of cost?",
      "En los últimos 12 meses, ¿hubo alguna vez que NECESITÓ atención pero NO LA RECIBIÓ por el costo?",
      "radio", options=YN_DK)

# ---------- Income, Birthplace, Address, Age ----------
add_q("Income","income_bracket",
      "What is your annual household income from all sources?",
      "¿Cuál es el ingreso anual total de su hogar (todas las fuentes)?",
      "radio", options=INCOME)
add_q("Income","num_supported",
      "How many people (including you) were supported by that income?",
      "¿Cuántas personas (incluyéndose) se mantuvieron con ese ingreso?",
      "int")
add_q("Birthplace","bp_where",
      "Where were you born?",
      "¿Dónde nació?",
      "radio", options=[opt("In the United States","En los Estados Unidos",1), opt("Outside the United States","Fuera de los Estados Unidos",2)])
add_q("Birthplace","bp_state",
      "If in the U.S., which state?",
      "Si fue en EE. UU., ¿en qué estado?",
      "text", branch=lambda a: a.get("bp_where",{}).get("code")==1)
add_q("Birthplace","bp_country",
      "If outside the U.S., which territory or country?",
      "Si fue fuera de EE. UU., ¿qué territorio o país?",
      "text", branch=lambda a: a.get("bp_where",{}).get("code")==2)

add_q("Address","addr_street","Street address (number and street) [optional]","Dirección (número y calle) [opcional]","text")
add_q("Address","addr_city","City","Ciudad","text")
add_q("Address","addr_state","State/Province","Estado/Provincia","text")
add_q("Address","addr_zip","ZIP/Postal code","Código postal","text")

add_q("Age","dob_mm","Birth month (MM)","Mes de nacimiento (MM)","text")
add_q("Age","dob_dd","Birth day (DD)","Día de nacimiento (DD)","text")
add_q("Age","dob_yy","Birth year (YYYY)","Año de nacimiento (AAAA)","text")
add_q("Age","age_years","Age (in years)","Edad (en años)","int")

# ---------- Employment / Marital / Education / English ----------
add_q("Employment","emp_status","Current employment status","Situación laboral actual","radio", options=EMP)
add_q("Employment","emp_other","If 'Other', please specify","Si seleccionó 'Otro', especifique","text", branch=lambda a: a.get("emp_status",{}).get("code")==8)
add_q("Marital Status","marital","Marital status","Estado civil","radio", options=MAR)
add_q("Education","education","Highest level of education completed","Máximo nivel educativo completado","radio", options=EDU)
add_q("English Proficiency","english_proficiency","How well do you speak English?","¿Qué tan bien habla inglés?","radio", options=ENG)

# ---------- Ethnicity & Race ----------
add_q("Ethnicity & Race","hispanic_origin","Are you of Hispanic/Latino/Spanish origin?","¿Es de origen hispano/latino/español?","radio",
      options=[
          opt("No, not of Hispanic/Latino/Spanish origin","No, no de origen hispano/latino/español",1),
          opt("Yes, Mexican/Mexican Am./Chicano","Sí, mexicano/mexicoamericano/chicano",2),
          opt("Yes, Puerto Rican","Sí, puertorriqueño",3),
          opt("Yes, Cuban","Sí, cubano",4),
          opt("Yes, another Hispanic/Latino/Spanish origin (specify)","Sí, otro origen hispano/latino/español (especifique)",5),
      ])
add_q("Ethnicity & Race","hispanic_origin_detail","Please specify your Hispanic/Latino/Spanish origin","Especifique su origen hispano/latino/español","text",
      branch=lambda a: a.get("hispanic_origin",{}).get("code")==5)

RACE_ITEMS = [
    ("race_white", "White (specify origins if you wish)","Blanco (especifique orígenes si desea)"),
    ("race_black", "Black or African American (specify origins if you wish)","Negro o afroamericano (especifique orígenes si desea)"),
    ("race_aian", "American Indian or Alaska Native (specify tribe)","Indígena americano o nativo de Alaska (especifique tribu)"),
    ("race_chinese", "Chinese","Chino"),
    ("race_filipino", "Filipino","Filipino"),
    ("race_asianind", "Asian Indian","Indio asiático"),
    ("race_viet", "Vietnamese","Vietnamita"),
    ("race_korean", "Korean","Coreano"),
    ("race_japanese", "Japanese","Japonés"),
    ("race_other_asian", "Other Asian (specify)","Otro asiático (especifique)"),
    ("race_nh", "Native Hawaiian","Nativo hawaiano"),
    ("race_samoan", "Samoan","Samoano"),
    ("race_chamorro", "Chamorro","Chamorro"),
    ("race_other_pi", "Other Pacific Islander (specify)","Otro isleño del Pacífico (especifique)"),
    ("race_other", "Some other race (specify)","Otra raza (especifique)"),
]
for rid, en, es in RACE_ITEMS:
    add_q("Ethnicity & Race", rid, f"Race — {en}", f"Raza — {es}", "radio",
          options=[opt("Select","Seleccionar",1), opt("Skip","Omitir",0)])
    if "specify" in en.lower() or "tribe" in en.lower():
        add_q("Ethnicity & Race", rid+"_detail", "Please specify", "Especifique", "text",
              branch=lambda a, key=rid: a.get(key,{}).get("code")==1)

# ---------- Food Insecurity ----------
add_q("Food Insecurity","fs1","“The food we bought just didn’t last, and we didn’t have money to get more.”",
      "“La comida que compramos no alcanzó, y no teníamos dinero para comprar más.”","radio", options=FS_OSN)
add_q("Food Insecurity","fs2","“We couldn’t afford to eat balanced meals.”",
      "“No podíamos permitirnos comer comidas balanceadas.”","radio", options=FS_OSN)
add_q("Food Insecurity","fs3","In the past 12 months, did you cut meal size or skip meals due to lack of money?",
      "En los últimos 12 meses, ¿recortó el tamaño de las comidas o se saltó comidas por falta de dinero?","radio", options=FS_YN)
add_q("Food Insecurity","fs3a","If yes, how often?","Si respondió sí, ¿con qué frecuencia?","radio", options=FS_3A, branch=lambda a: a.get("fs3",{}).get("code")==1)
add_q("Food Insecurity","fs4","In the past 12 months, did you eat less than you felt you should due to lack of money?",
      "En los últimos 12 meses, ¿comió menos de lo que debía por falta de dinero?","radio", options=FS_YN)
add_q("Food Insecurity","fs5","In the past 12 months, were you ever hungry but didn’t eat because there wasn’t enough money for food?",
      "En los últimos 12 meses, ¿tuvo hambre pero no comió por falta de dinero?","radio", options=FS_YN)

# ---------- Insurance ----------
INS_ITEMS = [
    ("ins_employer","Employer/union plan (yours/family; includes COBRA)","Plan de empleador/sindicato (suyo/familia; incluye COBRA)"),
    ("ins_direct","Direct purchase / Marketplace or exchange","Compra directa / Mercado o intercambio"),
    ("ins_medicare","Medicare","Medicare"),
    ("ins_medicaid","Medicaid / Medical Assistance / CHIP","Medicaid / Asistencia médica / CHIP"),
    ("ins_tricare","TRICARE / VA","TRICARE / VA"),
    ("ins_ihs","Indian Health Service","Servicio de Salud Indígena"),
    ("ins_other","Other health coverage","Otro tipo de cobertura de salud"),
]
for iid, en, es in INS_ITEMS:
    add_q("Health Insurance", iid, en, es, "radio",
          options=[opt("Covered","Con cobertura",1), opt("Not covered","Sin cobertura",2), opt("Not sure","No seguro(a)",3)])
add_q("Health Insurance","ins_other_detail","If 'Other' is covered, what type?","Si 'Otro' tiene cobertura, ¿qué tipo?","text",
      branch=lambda a: a.get("ins_other",{}).get("code")==1)

# ---------- Health Literacy ----------
add_q("Health Literacy","hl_conf","How confident are you filling out medical forms by yourself?",
      "¿Qué tan seguro(a) se siente al llenar formularios médicos por su cuenta?","radio", options=HL_CONF)
add_q("Health Literacy","hl_help_read","How often do you have someone help you read health materials?",
      "¿Con qué frecuencia alguien le ayuda a leer materiales de salud?","radio", options=FREQ5)
add_q("Health Literacy","hl_learn_prob","How often do you have problems learning about your medical condition due to written information?",
      "¿Con qué frecuencia tiene problemas para aprender sobre su condición por la información escrita?","radio", options=FREQ5)

# ---------- General Health & Sexual Orientation ----------
add_q("General Health","gen_health","Overall, how would you rate your health?","En general, ¿cómo califica su salud?","radio", options=GEN_HEALTH)
add_q("Sexual Orientation","so1","Which best represents how you think of yourself?","¿Cuál lo(a) representa mejor?","radio", options=SO1)
add_q("Sexual Orientation","so2","Additional options (if selected)","Opciones adicionales (si seleccionó)","radio", options=SO2, branch=lambda a: a.get("so1",{}).get("code")==5)
add_q("Sexual Orientation","so2_detail","Please describe","Describa","text", branch=lambda a: a.get("so2",{}).get("code")==10)

# ---------- Discrimination (Major) ----------
MAJ = [
    ("disc_fired","Unfairly fired from a job?","¿Despedido injustamente de un trabajo?"),
    ("disc_not_hired","Not hired for a job for unfair reasons?","¿No contratado por razones injustas?"),
    ("disc_denied_promo","Denied a promotion?","¿Negada una promoción?"),
    ("disc_police","Stopped/searched/threatened/abused by police?","¿Detenido/registrado/amenazado/abusado por la policía?"),
    ("disc_housing_blocked","Prevented from moving into a neighborhood?","¿Impedido de mudarse a un vecindario?"),
    ("disc_neighbors_hostile","Neighbors made life difficult after moving in?","¿Vecinos le hicieron la vida difícil tras mudarse?"),
    ("disc_bank_loan","Denied a bank loan?","¿Negado un préstamo bancario?"),
    ("disc_school_discouraged","Discouraged from continuing education?","¿Disuadido de continuar su educación?"),
    ("disc_healthcare","Denied/poorer healthcare than others?","¿Atención médica negada o de peor calidad que otros?"),
]
for k,en,es in MAJ:
    add_q("Discrimination (Major) ", k, en, es, "radio", options=YN)

# ---------- Discrimination (Everyday) ----------
EDS = [
    ("eds_courtesy","Treated with less courtesy than other people","Trato con menos cortesía que otras personas"),
    ("eds_respect","Treated with less respect than other people","Trato con menos respeto que otras personas"),
    ("eds_service","Received poorer service at restaurants or stores","Servicio peor en restaurantes o tiendas"),
    ("eds_not_smart","People act as if you are not smart","La gente actúa como si no fuera inteligente"),
    ("eds_afraid","People act as if they are afraid of you","La gente actúa como si le tuviera miedo"),
    ("eds_dishonest","People act as if you are dishonest","La gente actúa como si fuera deshonesto"),
    ("eds_not_as_good","People act as if you are not as good as they are","La gente actúa como si no fuera tan bueno como ellos"),
    ("eds_insulted","You are called names or insulted","Le ponen apodos o insultan"),
    ("eds_threatened","You are threatened or harassed","Le amenazan o acosan"),
]
for k,en,es in EDS:
    add_q("Discrimination (Everyday) ", k, en, es, "radio",
          options=[opt("Almost every day","Casi todos los días",1), opt("At least once a week","Al menos una vez por semana",2),
                   opt("A few times a month","Unas cuantas veces al mes",3), opt("A few times a year","Unas cuantas veces al año",4),
                   opt("Less than once a year","Menos de una vez al año",5), opt("Never","Nunca",6)])

# ---------- Neighborhood ----------
NBH = [
    ("nbh_safe_walk","I feel safe walking in my neighborhood, day or night.","Me siento seguro caminando en mi vecindario, de día o de noche."),
    ("nbh_violence","Violence is not a problem in my neighborhood.","La violencia no es un problema en mi vecindario."),
    ("nbh_safe_crime","My neighborhood is safe from crime.","Mi vecindario está a salvo del crimen."),
    ("nbh_active","I see people being active (walking, biking) in my neighborhood.","Veo gente activa (caminando, en bici) en mi vecindario."),
    ("nbh_sidewalks","There are sidewalks on most streets.","Hay aceras en la mayoría de las calles."),
    ("nbh_stores","There are stores within walking distance.","Hay tiendas a distancia a pie."),
    ("nbh_parks","There are parks or open spaces nearby.","Hay parques o espacios abiertos cerca."),
]
for k,en,es in NBH:
    add_q("Neighborhood", k, en, es, "radio", options=AGREE5)

# ---------- Housing ----------
add_q("Housing","house_bills","In the last 12 months, were you unable to pay mortgage/rent or utilities on time?",
      "En los últimos 12 meses, ¿no pudo pagar a tiempo la hipoteca/renta o los servicios?","radio", options=YN)
add_q("Housing","house_moves","How many times did you move in the past 12 months?","¿Cuántas veces se mudó en los últimos 12 meses?","int")
add_q("Housing","house_homeless","In the past 12 months, were you ever homeless (no stable place to live)?",
      "En los últimos 12 meses, ¿estuvo sin hogar (sin lugar estable para vivir)?","radio", options=YN)
add_q("Housing","house_sleep_place","If homeless, where did you sleep most often (shelter, street, car, doubled up, etc.)?",
      "Si estuvo sin hogar, ¿dónde durmió con más frecuencia (albergue, calle, auto, con conocidos, etc.)?","text",
      branch=lambda a: a.get("house_homeless",{}).get("code")==1)
add_q("Housing","house_forced_move","In the last 12 months, how many times were you forced to move by a landlord, bank, or mortgage company?",
      "En los últimos 12 meses, ¿cuántas veces fue obligado a mudarse por el propietario, banco o hipotecaria?","int")

# ---------- Transportation ----------
TRN = [
    ("trans_barrier","In the past 12 months, did lack of transportation keep you from appointments/work/essentials?",
     "En los últimos 12 meses, ¿la falta de transporte le impidió asistir a citas/trabajo/esenciales?"),
    ("trans_car_access","How often do you have access to a car when needed?",
     "¿Con qué frecuencia tiene acceso a un auto cuando lo necesita?"),
    ("trans_public_use","How often do you use public transportation?",
     "¿Con qué frecuencia usa transporte público?"),
    ("trans_public_reliable","How reliable is public transportation in your area?",
     "¿Qué tan confiable es el transporte público en su zona?"),
    ("trans_time_provider","How much time does it usually take to reach your healthcare provider?",
     "¿Cuánto tiempo suele tardar en llegar a su proveedor de salud?"),
    ("trans_cost_month","How much does transportation cost you per month (estimate)?",
     "¿Cuánto gasta en transporte por mes (estimado)?"),
    ("trans_rely_others","Do you rely on family/friends for rides?",
     "¿Depende de familia/amigos para traslados?"),
    ("trans_missed_essentials","Have you missed meds/food/essentials due to transport barriers?",
     "¿Ha dejado de obtener medicinas/comida/esenciales por barreras de transporte?"),
]
for k,en,es in TRN:
    if "cost" in k or "time" in k:
        add_q("Transportation", k, en, es, "int")
    else:
        add_q("Transportation", k, en, es, "radio", options=FREQ5 if "How often" in en else YN)

# ---------- Financial Strain ----------
add_q("Financial Strain","fin_bills_diff","How difficult is it for you to pay monthly bills?",
      "¿Qué tan difícil es pagar sus cuentas mensuales?","radio",
      options=[opt("Very difficult","Muy difícil",1), opt("Somewhat difficult","Algo difícil",2), opt("Not difficult","No es difícil",3)])
add_q("Financial Strain","fin_end_month","At the end of the month, how much money do you usually have?",
      "Al final del mes, ¿cuánto dinero suele tener?","radio",
      options=[opt("More than enough","Más que suficiente",1), opt("Just enough","Justo suficiente",2), opt("Not enough","No suficiente",3)])
add_q("Financial Strain","fin_utils_shut","In the past 12 months, were any utilities shut off due to nonpayment?",
      "En los últimos 12 meses, ¿algún servicio fue cortado por falta de pago?","radio", options=YN)
add_q("Financial Strain","fin_emergency_400","Do you have at least $400 available for an emergency?",
      "¿Tiene al menos $400 disponibles para una emergencia?","radio", options=YN)
add_q("Financial Strain","fin_payday_loans","Have you used payday loans/borrowed to cover living expenses?",
      "¿Usó préstamos de día de pago o pidió prestado para gastos de vida?","radio", options=YN)
add_q("Financial Strain","fin_help_others","Do you receive help from family/friends/community programs for living expenses?",
      "¿Recibe ayuda de familia/amigos/programas comunitarios para gastos de vida?","radio", options=YN)

# ---------- Social Support ----------
SS = [
    ("ss_listen","Someone is available to listen when you need to talk.","Alguien disponible para escuchar cuando necesita hablar."),
    ("ss_advice","Someone gives good advice when you are in trouble.","Alguien da buen consejo cuando tiene problemas."),
    ("ss_chores","Someone helps with chores if you are sick.","Alguien ayuda con tareas si está enfermo."),
    ("ss_emotional","Someone provides emotional support.","Alguien brinda apoyo emocional."),
    ("ss_private","Someone to share your private worries and fears.","Alguien con quien compartir preocupaciones privadas."),
    ("ss_love","Someone to love you and make you feel wanted.","Alguien que le ama y le hace sentir valorado."),
]
for k,en,es in SS:
    add_q("Social Support", k, en, es, "radio", options=FREQ5)

# ---------- Civic Engagement ----------
add_q("Civic Engagement","civ_registered","Are you registered to vote?","¿Está registrado para votar?","radio", options=YN)
add_q("Civic Engagement","civ_voted","Did you vote in the most recent national election?","¿Votó en la elección nacional más reciente?","radio", options=YN)
add_q("Civic Engagement","civ_volunteer","How often do you volunteer in your community?","¿Con qué frecuencia es voluntario en su comunidad?","radio", options=FREQ5)
add_q("Civic Engagement","civ_meetings","How often do you attend neighborhood/community meetings?","¿Con qué frecuencia asiste a reuniones comunitarias/vecinales?","radio", options=FREQ5)
add_q("Civic Engagement","civ_voice","Do you feel your voice is heard in local decision-making?","¿Siente que su voz se escucha en la toma de decisiones locales?","radio",
      options=[opt("Yes","Sí",1), opt("Sometimes","A veces",2), opt("No","No",3)])

# ---------- Work & Labor ----------
add_q("Work & Labor","work_sick_leave","Does your job provide paid sick leave?","¿Su trabajo ofrece licencia por enfermedad pagada?","radio", options=YN)
add_q("Work & Labor","work_insurance","Does your job provide health insurance?","¿Su trabajo ofrece seguro de salud?","radio", options=YN)
add_q("Work & Labor","work_retirement","Does your job provide retirement/pension benefits?","¿Su trabajo ofrece beneficios de jubilación/pensión?","radio", options=YN)
add_q("Work & Labor","work_min_wage","Does your job pay at least local minimum wage?","¿Su trabajo paga al menos el salario mínimo local?","radio", options=YN)
add_q("Work & Labor","work_union","Are you a member of a union/worker organization?","¿Es miembro de un sindicato/organización laboral?","radio", options=YN)

# ---------- Environment ----------
add_q("Environment","env_air_quality","Is the air quality in your neighborhood generally good, fair, or poor?",
      "¿La calidad del aire en su vecindario es buena, regular o mala?","radio",
      options=[opt("Good","Buena",1), opt("Fair","Regular",2), opt("Poor","Mala",3)])
add_q("Environment","env_exposure_health","Any health problems you think related to environmental exposures (pollution/lead/pesticides)?",
      "¿Problemas de salud que crea relacionados con exposiciones ambientales (contaminación/plomo/plaguicidas)?","radio", options=YN)
add_q("Environment","env_trash","How often do you see trash/litter/illegal dumping in your neighborhood?",
      "¿Con qué frecuencia ve basura/tiraderos ilegales en su vecindario?","radio", options=FREQ5)
add_q("Environment","env_outdoor_spaces","Do you have access to safe outdoor spaces for exercise and recreation?",
      "¿Tiene acceso a espacios exteriores seguros para ejercicio y recreación?","radio", options=YN)
add_q("Environment","env_safe_water","Do you have access to clean, safe tap water in your home?",
      "¿Tiene acceso a agua potable segura en su casa?","radio", options=YN)

# ---------- Community Resilience ----------
add_q("Community Resilience","res_safe_place_disaster","In a disaster (flood/fire/earthquake), do you have a safe place to go?",
      "En un desastre (inundación/incendio/terremoto), ¿tiene un lugar seguro a dónde ir?","radio", options=YN)
add_q("Community Resilience","res_comm_resources","Does your community have emergency preparedness resources?",
      "¿Su comunidad tiene recursos de preparación para emergencias?","radio", options=YN)
add_q("Community Resilience","res_info_access","Do you know how to access emergency information in your area?",
      "¿Sabe cómo acceder a la información de emergencias en su zona?","radio", options=YN)
add_q("Community Resilience","res_neighbors_help","Do you have neighbors/community you can rely on in an emergency?",
      "¿Tiene vecinos/comunidad en quienes confiar en una emergencia?","radio", options=YN)
add_q("Community Resilience","res_recover_quick","Do you feel your community could recover quickly after a disaster?",
      "¿Cree que su comunidad podría recuperarse rápidamente tras un desastre?","radio",
      options=[opt("Yes","Sí",1), opt("Maybe","Quizá",2), opt("No","No",3)])

# ---------- Tobacco ----------
add_q("Tobacco Use","tob_now_smoke","Do you now smoke cigarettes?","¿Actualmente fuma cigarrillos?","radio",
      options=[opt("Every day","Todos los días",1), opt("Some days","Algunos días",2), opt("Not at all","Nada",3)])
add_q("Tobacco Use","tob_100_cigs","Have you smoked at least 100 cigarettes in your life?",
      "¿Ha fumado al menos 100 cigarrillos en su vida?","radio", options=YN,
      branch=lambda a: a.get("tob_now_smoke",{}).get("code") in (1,2))
add_q("Tobacco Use","tob_age_start","At what age did you start smoking regularly?",
      "¿A qué edad comenzó a fumar regularmente?","int",
      branch=lambda a: a.get("tob_now_smoke",{}).get("code") in (1,2))
add_q("Tobacco Use","tob_quit_attempt","In the past 12 months, have you tried to quit?",
      "En los últimos 12 meses, ¿intentó dejar de fumar?","radio", options=YN,
      branch=lambda a: a.get("tob_now_smoke",{}).get("code") in (1,2))
add_q("Tobacco Use","tob_other_forms","Do you now use other tobacco (cigars/pipes/smokeless/e-cigarettes)?",
      "¿Usa ahora otros productos de tabaco (puros/pipas/tabaco sin humo/cigarrillos electrónicos)?","radio", options=YN)

# ---------- Alcohol ----------
add_q("Alcohol Use","alc_ever","Have you ever had a drink of any kind of alcoholic beverage?",
      "¿Alguna vez ha tomado alguna bebida alcohólica?","radio", options=YN)
add_q("Alcohol Use","alc_30d_any","During the past 30 days, did you have at least one drink of alcohol?",
      "En los últimos 30 días, ¿tomó al menos una bebida alcohólica?","radio", options=YN,
      branch=lambda a: a.get("alc_ever",{}).get("code")==1)
add_q("Alcohol Use","alc_30d_days","During the past 30 days, on how many days did you drink alcohol?",
      "En los últimos 30 días, ¿en cuántos días bebió alcohol?","int",
      branch=lambda a: a.get("alc_30d_any",{}).get("code")==1)
add_q("Alcohol Use","alc_30d_usual","On drinking days, how many drinks did you usually have?",
      "En días que bebió, ¿cuántas bebidas suele tomar?","int",
      branch=lambda a: a.get("alc_30d_any",{}).get("code")==1)
add_q("Alcohol Use","alc_30d_max","During the past 30 days, what is the largest number of drinks you had on any occasion?",
      "En los últimos 30 días, ¿cuál fue el mayor número de bebidas en una ocasión?","int",
      branch=lambda a: a.get("alc_30d_any",{}).get("code")==1)
add_q("Alcohol Use","alc_binge_year","In the past year, how often did you have 5 (men)/4 (women) or more drinks on one occasion?",
      "En el último año, ¿con qué frecuencia tomó 5 (hombres)/4 (mujeres) o más bebidas en una ocasión?","radio",
      options=[opt("Never","Nunca",0), opt("Less than monthly","Menos que mensual",1), opt("Monthly","Mensual",2), opt("Weekly","Semanal",3), opt("Daily or almost daily","Diaria o casi diaria",4)],
      branch=lambda a: a.get("alc_ever",{}).get("code")==1)
for k,en,es in [
    ("alc_cut","Have you felt you should cut down on drinking?","¿Ha sentido que debería reducir su consumo?"),
    ("alc_annoy","Have people annoyed you by criticizing your drinking?","¿Le han molestado al criticar su consumo?"),
    ("alc_guilt","Have you felt bad or guilty about your drinking?","¿Se ha sentido mal o culpable por su consumo?"),
    ("alc_eyeopener","Have you had a drink first thing in the morning (eye-opener)?","¿Ha bebido al despertar (para calmar los nervios)?"),
]:
    add_q("Alcohol Use", k, en, es, "radio", options=YN, branch=lambda a: a.get("alc_ever",{}).get("code")==1)

# ---------- Digital Access ----------
add_q("Digital Access","net_home","Do you have Internet access at home?","¿Tiene acceso a Internet en casa?","radio", options=YN)
add_q("Digital Access","smartphone","Do you have a smartphone or tablet with Internet?","¿Tiene un teléfono inteligente o tableta con Internet?","radio", options=YN)
add_q("Digital Access","net_health_info","How often do you look up health information online?","¿Con qué frecuencia busca información de salud en línea?","radio", options=FREQ5)
add_q("Digital Access","net_confidence","How confident are you finding helpful health resources online?","¿Qué tan seguro(a) se siente para encontrar recursos de salud útiles en línea?","radio", options=HL_CONF)
add_q("Digital Access","portal_comm","Do you use email/patient portals to communicate with your clinic?","¿Usa correo o portales del paciente para comunicarse con su clínica?","radio", options=YN)

# =========================
# Streamlit UI
# =========================
#st.set_page_config(page_title="SDoH Survey", page_icon="🏥", layout="wide")

st.set_page_config(
    page_title="SDoH Survey",
    page_icon="🏥",
    layout="wide",
    initial_sidebar_state="collapsed"  # 👈 hides sidebar by default
)

# Add this near the top of your app (right after st.set_page_config)
st.markdown("""
<script>
document.addEventListener('keydown', function(e) {
    // Ctrl + D or Command + D
    if ((e.ctrlKey || e.metaKey) && e.key === 'd') {
        const sidebar = window.parent.document.querySelector('section[data-testid="stSidebar"]');
        if (sidebar) {
            sidebar.style.display = (sidebar.style.display === 'none' || sidebar.style.display === '') ? 'block' : 'none';
        }
    }
});
</script>
""", unsafe_allow_html=True)


# Style: modern headers, rounded buttons
st.markdown("""
<style>
h3 { font-size: 1.6rem !important; }
.stRadio > div { gap: .5rem; }
div.stButton > button { border-radius: 999px; height: 46px; font-weight: 600; }
.badge { display:inline-block; padding:3px 10px; border-radius:999px; background:#eef5ff; color:#246; font-size:.85rem; margin-bottom:6px; }
.card { background:white; border-radius:16px; padding:18px 20px; box-shadow: 0 4px 18px rgba(0,0,0,.08); }
</style>
""", unsafe_allow_html=True)

st.markdown("<div style='text-align:center' class='card'><h2 style='color:#0a5bd6;margin:.2rem 0'>🏥 Social Determinants of Health — Patient Form</h2><div>Welcome! / ¡Bienvenido!</div></div>", unsafe_allow_html=True)
st.write("")

# Language choice (centered)
col_a, col_b, col_c = st.columns([1,2,1])
with col_b:
    lang_choice = st.radio("Language / Idioma", ["English", "Español"], horizontal=True, index=0)
lang = "en" if lang_choice == "English" else "es"

# ---- Radio helper with NO default ----
def radio_force_click(label, options_labels, key):
    # index=None => no default; requires Streamlit >= 1.25
    picked = st.radio(label, options_labels, index=None, key=f"radio_{key}", label_visibility="visible")
    return picked if picked is not None else None

# Branching helper (safe)
def is_visible(q, ans_dict):
    br = q.get("branch")
    if br is None:
        return True
    try:
        return bool(br(ans_dict))
    except Exception:
        return True

def fs_category(answers_dict):
    def code(k):
        v = answers_dict.get(k)
        if isinstance(v, dict):
            return v.get("code")
        return v
    affirm = 0
    if code("fs1") in (1,2): affirm += 1
    if code("fs2") in (1,2): affirm += 1
    if code("fs3")==1:
        affirm += 1
        if code("fs3a") in (1,2): affirm += 1
    if code("fs4")==1: affirm += 1
    if code("fs5")==1: affirm += 1
    if affirm<=1: cat="High or marginal food security"
    elif affirm<=4: cat="Low food security"
    else: cat="Very low food security"
    return affirm, cat

# Prepass snapshot from session_state (for section visibility)
def snapshot_from_state():
    snap = {}
    for q in QUESTIONS:
        if q["type"] == "radio":
            lbl = st.session_state.get(f"radio_{q['id']}", None)
            if lbl is None:
                snap[q["id"]] = {"code": None, "label": None}
            else:
                code = next((o["code"] for o in q["options"] if o[lang] == lbl), None)
                snap[q["id"]] = {"code": code, "label": lbl}
        elif q["type"] == "int":
            snap[q["id"]] = st.session_state.get(f"num_{q['id']}", 0)
        else:
            snap[q["id"]] = st.session_state.get(f"text_{q['id']}", "").strip()
    return snap

answers_snapshot = snapshot_from_state()

# Determine which sections have any visible questions
seen_sections = []
section_visible = {}
for q in QUESTIONS:
    if q["section"] not in seen_sections:
        seen_sections.append(q["section"])
        section_visible[q["section"]] = False
    if is_visible(q, answers_snapshot):
        section_visible[q["section"]] = True

# =========================
# Render Sections as Compact Expanders with Counts + Indented Branch Questions
# =========================

SECTION_TITLES_ES = {
    "Access to Health Services": "Acceso a los servicios de salud",
    "Income": "Ingresos",
    "Birthplace": "Lugar de nacimiento",
    "Address": "Dirección",
    "Age": "Edad",
    "Employment": "Empleo",
    "Marital Status": "Estado civil",
    "Education": "Educación",
    "English Proficiency": "Dominio del inglés",
    "Ethnicity & Race": "Etnicidad y raza",
    "Food Insecurity": "Inseguridad alimentaria",
    "Health Insurance": "Seguro de salud",
    "Health Literacy": "Alfabetización en salud",
    "General Health": "Salud general",
    "Sexual Orientation": "Orientación sexual",
    "Discrimination (Major)": "Discriminación (mayor)",
    "Discrimination (Everyday)": "Discriminación (cotidiana)",
    "Neighborhood": "Vecindario",
    "Housing": "Vivienda",
    "Transportation": "Transporte",
    "Financial Strain": "Dificultades financieras",
    "Social Support": "Apoyo social",
    "Civic Engagement": "Participación cívica",
    "Work & Labor": "Trabajo y empleo",
    "Environment": "Medio ambiente",
    "Community Resilience": "Resiliencia comunitaria",
    "Tobacco Use": "Consumo de tabaco",
    "Alcohol Use": "Consumo de alcohol",
    "Digital Access": "Acceso digital"
}

# Compact styling and indentation for sub-questions
st.markdown("""
<style>
    div.block-container {padding-top: 1rem !important;}
    .stExpander {margin-bottom: 4px !important;}
    .stExpander div[role="button"] {padding: 0.3rem 0.6rem !important;}
    .stExpanderContent {padding-top: 0.4rem !important; padding-bottom: 0.4rem !important;}
    h3, h4, h5, p {margin: 0.15rem 0 !important;}
    .subquestion {margin-left: 25px !important; border-left: 2px solid #d0d7de; padding-left: 10px; background: #fafafa; border-radius: 4px;}
</style>
""", unsafe_allow_html=True)

answers = {}
qnum = 1

for section in seen_sections:
    if not section_visible.get(section):
        continue

    visible_qs = [q for q in QUESTIONS if q["section"] == section and is_visible(q, answers_snapshot)]
    q_count = len(visible_qs)
    section_label = section if lang == "en" else SECTION_TITLES_ES.get(section, section)
    label_text = (
        f"📂 {section_label} — {q_count} question{'s' if q_count != 1 else ''}"
        if lang == "en" else
        f"📂 {section_label} — {q_count} pregunta{'s' if q_count != 1 else ''}"
    )

    with st.expander(label_text, expanded=False):
        for q in [qq for qq in QUESTIONS if qq["section"] == section]:
            if not is_visible(q, answers):
                continue

            label_txt = q["text"][lang]

            # Determine if it's a sub-question (i.e., has a branch condition)
            if q.get("branch") is not None:
                st.markdown(f"<div class='subquestion'><strong>{qnum}) {label_txt}</strong>", unsafe_allow_html=True)
            else:
                st.markdown(f"<p style='margin-bottom:2px;'><strong>{qnum}) {label_txt}</strong></p>", unsafe_allow_html=True)

            # Render question input (inside subquestion div if applicable)
            container = st.container()
            with container:
                if q["type"] == "radio":
                    opts = q["options"]
                    labels_local = [o[lang] for o in opts]
                    picked_label = radio_force_click("", labels_local, key=q["id"])
                    if picked_label is None:
                        answers[q["id"]] = {"code": None, "label": None}
                    else:
                        code = next((o["code"] for o in opts if o[lang] == picked_label), None)
                        answers[q["id"]] = {"code": code, "label": picked_label}

                elif q["type"] == "int":
                    v = st.number_input("", min_value=0, step=1, key=f"num_{q['id']}")
                    answers[q["id"]] = v

                else:  # text input
                    v = st.text_input("", key=f"text_{q['id']}")
                    answers[q["id"]] = v.strip()

            if q.get("branch") is not None:
                st.markdown("</div>", unsafe_allow_html=True)

            qnum += 1


# Submit centered
col1, col2, col3 = st.columns([1,2,1])
with col2:
    if st.button("✅ Submit" if lang=="en" else "✅ Enviar", use_container_width=True):
        raw,cat = fs_category(answers)
        record = {
            "meta":{"completed_at":now_iso(),"instrument":"SDoH Bilingual Full (Streamlit) v1.0","language":lang},
            "sections":{},
            "derived":{"food_security_raw_score": raw, "food_security_category": cat}
        }
        for q in QUESTIONS:
            sec = q["section"]
            record["sections"].setdefault(sec, {})
            if q["id"] in answers:
                record["sections"][sec][q["id"]] = answers[q["id"]]

        save_all_outputs(record)
        st.success("✅ Thank you! Survey complete." if lang=="en" else "✅ ¡Gracias! Encuesta completada.")
        st.balloons()














