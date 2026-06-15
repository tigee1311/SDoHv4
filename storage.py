"""Hospital-isolated Excel persistence for the SDoH Streamlit survey.

This workbook-based layer is intended as a practical first persistence step for
research/testing deployments. For true HIPAA or clinical production use, move
this data into an authenticated, encrypted database with role-based access
control, audit logs, backups, retention policies, and a formal risk review.
"""

from __future__ import annotations

import datetime as dt
import json
import os
import re
import tempfile
import uuid
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook


RESPONSE_WORKBOOK = Path(os.getenv("SDOH_RESPONSE_WORKBOOK", "sdoh_responses.xlsx"))
HOSPITAL_INDEX_SHEET = "_hospitals"
SECTION_SETTINGS_SHEET = "_hospital_sections"

HOSPITAL_INDEX_COLUMNS = ["hospital_name", "sheet_name", "created_at"]
SECTION_SETTINGS_COLUMNS = ["hospital_name", "updated_at", "enabled_sections_json"]
RESPONSE_COLUMNS = [
    "timestamp",
    "hospital_name",
    "session_id",
    "save_id",
    "status",
    "completion_percentage",
    "language",
    "instrument",
    "category",
    "question_id",
    "question_text_en",
    "question_text_es",
    "response",
    "response_code",
]


def now_iso() -> str:
    """Return a second-resolution local timestamp for response rows."""
    return dt.datetime.now().isoformat(timespec="seconds")


def sanitize_sheet_name(hospital_name: str) -> str:
    """Return an Excel-safe sheet name derived from a hospital name."""
    cleaned = re.sub(r"[\[\]\:\*\?\/\\]", "", hospital_name or "").strip()
    cleaned = re.sub(r"\s+", "_", cleaned).strip("'")
    if not cleaned:
        cleaned = "Hospital"
    return cleaned[:31]


def get_existing_hospitals(workbook_path: str | Path = RESPONSE_WORKBOOK) -> list[str]:
    """Return configured hospital display names without exposing response data."""
    path = Path(workbook_path)
    if not path.exists():
        return []

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return []

    try:
        if HOSPITAL_INDEX_SHEET in wb.sheetnames:
            ws = wb[HOSPITAL_INDEX_SHEET]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return []
            header = [str(value) if value is not None else "" for value in rows[0]]
            if "hospital_name" not in header:
                return []
            name_idx = header.index("hospital_name")
            hospitals = [
                str(row[name_idx]).strip()
                for row in rows[1:]
                if len(row) > name_idx and row[name_idx] not in (None, "")
            ]
            return sorted(dict.fromkeys(hospitals), key=str.casefold)

        # Backward-compatible inference for workbooks already using hospital tabs.
        hospitals = []
        for sheet_name in wb.sheetnames:
            if sheet_name.startswith("_") or sheet_name == "responses":
                continue
            ws = wb[sheet_name]
            header = [cell.value for cell in next(ws.iter_rows(max_row=1), [])]
            if "hospital_name" in header:
                name_idx = header.index("hospital_name")
                first_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), None)
                if first_row and len(first_row) > name_idx and first_row[name_idx]:
                    hospitals.append(str(first_row[name_idx]).strip())
                else:
                    hospitals.append(sheet_name)
        return sorted(dict.fromkeys(hospitals), key=str.casefold)
    finally:
        wb.close()


def ensure_hospital_sheet(
    hospital_name: str,
    workbook_path: str | Path = RESPONSE_WORKBOOK,
) -> str:
    """Create or return the hospital's isolated worksheet name."""
    display_name = _normalize_hospital_name(hospital_name)
    path = Path(workbook_path)
    wb = _load_or_create_workbook(path)
    try:
        sheet_name = _ensure_hospital_sheet_in_workbook(wb, display_name)
        _atomic_save_workbook(wb, path)
        return sheet_name
    finally:
        wb.close()


def save_responses(
    hospital_name: str,
    session_id: str,
    responses: dict[str, Any],
    status: str,
    *,
    questions: list[dict[str, Any]] | None = None,
    completion_percentage: int | float = 0,
    language: str = "en",
    instrument: str = "SDoH Bilingual Full (Streamlit) v1.0",
    derived: dict[str, Any] | None = None,
    workbook_path: str | Path = RESPONSE_WORKBOOK,
) -> dict[str, Any]:
    """Append one partial or completed save event to the hospital worksheet."""
    if status not in {"partial", "completed"}:
        raise ValueError("status must be 'partial' or 'completed'")

    display_name = _normalize_hospital_name(hospital_name)
    if not session_id:
        raise ValueError("session_id is required")

    path = Path(workbook_path)
    wb = _load_or_create_workbook(path)
    timestamp = now_iso()
    save_id = str(uuid.uuid4())

    try:
        sheet_name = _ensure_hospital_sheet_in_workbook(wb, display_name)
        ws = wb[sheet_name]
        header = _ensure_response_headers(ws)
        rows = _response_rows(
            timestamp=timestamp,
            hospital_name=display_name,
            session_id=session_id,
            save_id=save_id,
            status=status,
            completion_percentage=completion_percentage,
            language=language,
            instrument=instrument,
            responses=responses,
            questions=questions,
            derived=derived or {},
        )

        for row in rows:
            ws.append([row.get(column, "") for column in header])

        _atomic_save_workbook(wb, path)
        return {
            "workbook_path": str(path),
            "sheet_name": sheet_name,
            "save_id": save_id,
            "rows_written": len(rows),
        }
    finally:
        wb.close()


def load_hospital_data(
    hospital_name: str,
    workbook_path: str | Path = RESPONSE_WORKBOOK,
) -> pd.DataFrame:
    """Load only the selected hospital's worksheet as a DataFrame."""
    path = Path(workbook_path)
    if not path.exists():
        return pd.DataFrame(columns=RESPONSE_COLUMNS)

    sheet_name = _find_hospital_sheet(path, hospital_name)
    if not sheet_name:
        return pd.DataFrame(columns=RESPONSE_COLUMNS)

    try:
        return pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")
    except ValueError:
        return pd.DataFrame(columns=RESPONSE_COLUMNS)


def export_hospital_workbook_bytes(
    hospital_name: str,
    workbook_path: str | Path = RESPONSE_WORKBOOK,
) -> bytes:
    """Build a single-hospital Excel export for UI downloads."""
    df = load_hospital_data(hospital_name, workbook_path=workbook_path)
    output = BytesIO()
    sheet_name = sanitize_sheet_name(hospital_name) or "responses"
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    output.seek(0)
    return output.getvalue()


def load_hospital_section_settings(
    hospital_name: str,
    workbook_path: str | Path = RESPONSE_WORKBOOK,
) -> list[str] | None:
    """Return saved enabled survey sections for a hospital, or None for default-all."""
    path = Path(workbook_path)
    if not path.exists():
        return None

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None

    try:
        if SECTION_SETTINGS_SHEET not in wb.sheetnames:
            return None

        ws = wb[SECTION_SETTINGS_SHEET]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return None

        header = [str(value) if value is not None else "" for value in rows[0]]
        if "hospital_name" not in header or "enabled_sections_json" not in header:
            return None

        name_idx = header.index("hospital_name")
        sections_idx = header.index("enabled_sections_json")
        display_name = _normalize_hospital_name(hospital_name)
        for row in rows[1:]:
            if len(row) <= max(name_idx, sections_idx):
                continue
            saved_name = "" if row[name_idx] is None else str(row[name_idx]).strip()
            if saved_name.casefold() != display_name.casefold():
                continue
            raw_sections = row[sections_idx]
            if raw_sections in (None, ""):
                return None
            try:
                sections = json.loads(str(raw_sections))
            except json.JSONDecodeError:
                return None
            if not isinstance(sections, list):
                return None
            return [str(section).strip() for section in sections if str(section).strip()]

        return None
    finally:
        wb.close()


def save_hospital_section_settings(
    hospital_name: str,
    enabled_sections: list[str],
    workbook_path: str | Path = RESPONSE_WORKBOOK,
) -> dict[str, Any]:
    """Persist enabled survey sections for the selected hospital."""
    display_name = _normalize_hospital_name(hospital_name)
    sections = [str(section).strip() for section in enabled_sections if str(section).strip()]
    if not sections:
        raise ValueError("At least one section must be selected")

    path = Path(workbook_path)
    wb = _load_or_create_workbook(path)
    timestamp = now_iso()

    try:
        _ensure_hospital_sheet_in_workbook(wb, display_name)
        ws = _ensure_section_settings_sheet(wb)
        header = _ensure_headers(ws, SECTION_SETTINGS_COLUMNS)
        name_idx = header.index("hospital_name") + 1
        updated_idx = header.index("updated_at") + 1
        sections_idx = header.index("enabled_sections_json") + 1

        saved = False
        encoded_sections = json.dumps(sections, ensure_ascii=False)
        for row_idx in range(2, ws.max_row + 1):
            saved_name = ws.cell(row=row_idx, column=name_idx).value
            saved_name = "" if saved_name is None else str(saved_name).strip()
            if saved_name.casefold() == display_name.casefold():
                ws.cell(row=row_idx, column=updated_idx, value=timestamp)
                ws.cell(row=row_idx, column=sections_idx, value=encoded_sections)
                saved = True
                break

        if not saved:
            row = {column: "" for column in header}
            row["hospital_name"] = display_name
            row["updated_at"] = timestamp
            row["enabled_sections_json"] = encoded_sections
            ws.append([row.get(column, "") for column in header])

        _atomic_save_workbook(wb, path)
        return {
            "workbook_path": str(path),
            "hospital_name": display_name,
            "updated_at": timestamp,
            "enabled_sections": sections,
        }
    finally:
        wb.close()


def _normalize_hospital_name(hospital_name: str) -> str:
    name = re.sub(r"\s+", " ", hospital_name or "").strip()
    if not name:
        raise ValueError("Hospital name is required")
    return name


def _load_or_create_workbook(path: Path):
    if path.exists():
        return load_workbook(path)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = HOSPITAL_INDEX_SHEET
    ws.append(HOSPITAL_INDEX_COLUMNS)
    return wb


def _ensure_hospital_index(wb) -> Any:
    if HOSPITAL_INDEX_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(HOSPITAL_INDEX_SHEET, 0)
        ws.append(HOSPITAL_INDEX_COLUMNS)
        return ws

    ws = wb[HOSPITAL_INDEX_SHEET]
    _ensure_headers(ws, HOSPITAL_INDEX_COLUMNS)
    return ws


def _ensure_section_settings_sheet(wb) -> Any:
    if SECTION_SETTINGS_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(SECTION_SETTINGS_SHEET)
        ws.append(SECTION_SETTINGS_COLUMNS)
        return ws

    ws = wb[SECTION_SETTINGS_SHEET]
    _ensure_headers(ws, SECTION_SETTINGS_COLUMNS)
    return ws


def _ensure_hospital_sheet_in_workbook(wb, hospital_name: str) -> str:
    index_ws = _ensure_hospital_index(wb)
    records = _hospital_index_records(index_ws)
    matching = next(
        (
            record
            for record in records
            if record["hospital_name"].strip().casefold() == hospital_name.casefold()
        ),
        None,
    )

    if matching:
        sheet_name = matching["sheet_name"]
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            _ensure_response_headers(ws)
        else:
            _ensure_response_headers(wb[sheet_name])
        return sheet_name

    sheet_name = _unique_sheet_name(sanitize_sheet_name(hospital_name), wb.sheetnames)
    ws = wb.create_sheet(sheet_name)
    _ensure_response_headers(ws)
    index_ws.append([hospital_name, sheet_name, now_iso()])
    return sheet_name


def _hospital_index_records(index_ws) -> list[dict[str, str]]:
    header = [cell.value for cell in index_ws[1]]
    records = []
    for row in index_ws.iter_rows(min_row=2, values_only=True):
        record = {
            column: str(row[idx]).strip()
            for idx, column in enumerate(header)
            if column and idx < len(row) and row[idx] not in (None, "")
        }
        if record.get("hospital_name") and record.get("sheet_name"):
            records.append(record)
    return records


def _unique_sheet_name(base_name: str, existing_names: list[str]) -> str:
    existing = {name.casefold() for name in existing_names}
    candidate = base_name[:31]
    if candidate.casefold() not in existing and candidate != HOSPITAL_INDEX_SHEET:
        return candidate

    for index in range(2, 1000):
        suffix = f"_{index}"
        candidate = f"{base_name[:31 - len(suffix)]}{suffix}"
        if candidate.casefold() not in existing and candidate != HOSPITAL_INDEX_SHEET:
            return candidate

    raise RuntimeError("Could not create a unique Excel sheet name")


def _ensure_headers(ws, expected_columns: list[str]) -> list[str]:
    if ws.max_row == 0:
        ws.append(expected_columns)
        return expected_columns

    existing = [cell.value for cell in ws[1]]
    if not any(existing):
        for col_idx, column in enumerate(expected_columns, start=1):
            ws.cell(row=1, column=col_idx, value=column)
        return expected_columns

    header = [str(value) if value is not None else "" for value in existing]
    for column in expected_columns:
        if column not in header:
            header.append(column)
            ws.cell(row=1, column=len(header), value=column)
    return header


def _ensure_response_headers(ws) -> list[str]:
    return _ensure_headers(ws, RESPONSE_COLUMNS)


def _response_rows(
    *,
    timestamp: str,
    hospital_name: str,
    session_id: str,
    save_id: str,
    status: str,
    completion_percentage: int | float,
    language: str,
    instrument: str,
    responses: dict[str, Any],
    questions: list[dict[str, Any]] | None,
    derived: dict[str, Any],
) -> list[dict[str, Any]]:
    question_rows = questions or [
        {"id": question_id, "section": "", "text": {"en": question_id, "es": question_id}}
        for question_id in responses
    ]

    rows = [
        _base_response_row(
            timestamp=timestamp,
            hospital_name=hospital_name,
            session_id=session_id,
            save_id=save_id,
            status=status,
            completion_percentage=completion_percentage,
            language=language,
            instrument=instrument,
            category=question.get("section", ""),
            question_id=question["id"],
            question_text_en=question.get("text", {}).get("en", ""),
            question_text_es=question.get("text", {}).get("es", ""),
            answer=responses.get(question["id"]),
        )
        for question in question_rows
    ]

    for key, value in derived.items():
        rows.append(
            _base_response_row(
                timestamp=timestamp,
                hospital_name=hospital_name,
                session_id=session_id,
                save_id=save_id,
                status=status,
                completion_percentage=completion_percentage,
                language=language,
                instrument=instrument,
                category="Derived",
                question_id=key,
                question_text_en=key.replace("_", " ").title(),
                question_text_es=key.replace("_", " ").title(),
                answer=value,
            )
        )

    return rows


def _base_response_row(
    *,
    timestamp: str,
    hospital_name: str,
    session_id: str,
    save_id: str,
    status: str,
    completion_percentage: int | float,
    language: str,
    instrument: str,
    category: str,
    question_id: str,
    question_text_en: str,
    question_text_es: str,
    answer: Any,
) -> dict[str, Any]:
    response, response_code = _serialize_answer(answer)
    return {
        "timestamp": timestamp,
        "hospital_name": hospital_name,
        "session_id": session_id,
        "save_id": save_id,
        "status": status,
        "completion_percentage": round(float(completion_percentage), 2),
        "language": language,
        "instrument": instrument,
        "category": category,
        "question_id": question_id,
        "question_text_en": question_text_en,
        "question_text_es": question_text_es,
        "response": response,
        "response_code": response_code,
    }


def _serialize_answer(answer: Any) -> tuple[str, str]:
    if answer is None:
        return "", ""
    if isinstance(answer, dict):
        label = answer.get("label")
        code = answer.get("code")
        return "" if label is None else str(label), "" if code is None else str(code)
    if isinstance(answer, (list, tuple)):
        return json.dumps(answer, ensure_ascii=False), ""
    return str(answer), ""


def _find_hospital_sheet(path: Path, hospital_name: str) -> str | None:
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None

    try:
        display_name = _normalize_hospital_name(hospital_name)
        if HOSPITAL_INDEX_SHEET in wb.sheetnames:
            ws = wb[HOSPITAL_INDEX_SHEET]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                header = [str(value) if value is not None else "" for value in rows[0]]
                if "hospital_name" in header and "sheet_name" in header:
                    name_idx = header.index("hospital_name")
                    sheet_idx = header.index("sheet_name")
                    for row in rows[1:]:
                        if len(row) <= max(name_idx, sheet_idx):
                            continue
                        saved_name = "" if row[name_idx] is None else str(row[name_idx]).strip()
                        saved_sheet = "" if row[sheet_idx] is None else str(row[sheet_idx]).strip()
                        if saved_name.casefold() == display_name.casefold() and saved_sheet in wb.sheetnames:
                            return saved_sheet

        sheet_name = sanitize_sheet_name(display_name)
        return sheet_name if sheet_name in wb.sheetnames else None
    finally:
        wb.close()


def _atomic_save_workbook(wb, path: Path) -> None:
    """Save to a temporary workbook and replace the target to reduce corruption risk."""
    path.parent.mkdir(parents=True, exist_ok=True)
    handle, temp_name = tempfile.mkstemp(
        prefix=f".{path.name}.",
        suffix=".tmp.xlsx",
        dir=str(path.parent),
    )
    os.close(handle)
    temp_path = Path(temp_name)

    try:
        wb.save(temp_path)
        os.replace(temp_path, path)
    except PermissionError as exc:
        raise RuntimeError(
            "Could not save responses. Close the Excel workbook if it is open and try again."
        ) from exc
    except Exception as exc:
        raise RuntimeError("Could not save responses without risking workbook corruption.") from exc
    finally:
        if temp_path.exists():
            temp_path.unlink()
